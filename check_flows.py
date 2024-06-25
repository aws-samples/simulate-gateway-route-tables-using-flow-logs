#!/usr/bin/env python3
# Script to check the aggregated VPC Flog summaries with a network configuration CSV
# See README.md for more information.

import os
import argparse
import sys
import ipaddress
import csv
from openpyxl import Workbook
from openpyxl import load_workbook


def excel_to_cidr_dict(excel_filename, show_warnings=False):
  '''
  Takes an Excel file with the following columns:
      'cidr','account_id','name','vpc_id','associate_with','propagate_to'
  and converts it into a hash map (dictionary) of /24 CIDRs -> network info dict.
  It is converted to /24 blocks because the Athena CSV file is uses /24 blocks.

  Sample input Excel:
    10.123.0.0/23 | 012345678912 | Some VPC | vpc-abcd12345 | Isolated | Infrastructure,Onpremises
  Sample output:
  {
    '10.123.0' : {
      'cidr': '10.123.0.0/23',
      'account_id': '000000000000',
      'name': 'Some VPC',
      'vpc_id': 'vpc-abcd12345',
      'associate_with': 'Isolated',
      'propagate_to': [
        'Infrastructure',
        'Onpremises'
      ]
    },
    '10.123.1' : {
      'cidr': '10.123.0.0/23',
      'account_id': '000000000000',
      'name': 'Some VPC',
      'vpc_id': 'vpc-abcd12345',
      'associate_with': 'Isolated',
      'propagate_to': [
        'Infrastructure',
        'Onpremises'
      ]
    }    
  }
  '''

  return_dict = {}

  # Iterate through Excel file:
  header = [ 'cidr','account_id','name','vpc_id','associate_with','propagate_to' ]
  wb = load_workbook(filename=excel_filename, read_only=True)
  sheet = wb.active
  for row in sheet.iter_rows(min_row=2):
    cidr = row[header.index('cidr')].value
    # Remove spaces, for example "10.215.96.0 /21" -> "10.215.96.0/21"
    cidr = cidr.replace(' ', '')
    account_id = row[header.index('account_id')].value
    name = row[header.index('name')].value
    vpc_id = row[header.index('vpc_id')].value
    associate_with = row[header.index('associate_with')].value
    propagate_to_str = row[header.index('propagate_to')].value

    if propagate_to_str:
      propagate_to = [ x.strip() for x in propagate_to_str.split(',') ]
    else:
      propagate_to = []

    # The memory required for the info dict is shared for a CIDR split into several /24 blocks, as it's a pointer:
    information = {
        'cidr': cidr,
        'account_id': account_id,
        'name': name,
        'vpc_id': vpc_id,
        'associate_with': associate_with,
        'propagate_to': propagate_to
    }

    # Split the CIDR into /24 blocks:
    try:
      for cidr24 in ipaddress.ip_network(cidr).subnets(new_prefix=24):
        cidr24_string = cidr24.exploded
        # Convert 10.123.4.0/24 -> 10.123.4
        # This is to match the columns in the Athena CSV output.
        key = ".".join(cidr24_string.split('.')[:3])

        return_dict[key] = information
    except ValueError:
      if show_warnings:
        print(f'Skipping {cidr} ({name}), shorter than /24. ')

  wb.close()
  return return_dict
  
def main():
  # Process arguments:
  parser = argparse.ArgumentParser()
  parser.add_argument('--flowlogcsv', required=True,
                      help='Path to the Athena VPC flow log query CSV')
  parser.add_argument('--xlsx', required=True,
                      help='Path to the network CIDR/VPC associations/propagations mapping XLSX')
  parser.add_argument('--show-warnings', action='store_true', help='Show unmatched CIDRs')
                      
  args = parser.parse_args()

  xlsx_filename = args.xlsx
  flowlog_filename = args.flowlogcsv
  show_warnings = args.show_warnings
  for filename in (xlsx_filename, flowlog_filename):
    if not os.path.isfile(filename):
      print(f'{filename} does not exist.')
      sys.exit(1)

  # Use this set() to deduplicate the output strings, as there are many /24s for each VPC CIDR
  failed_strings = set()
  warning_strings = set()

  # Get 24/CIDR to information dict:
  info = excel_to_cidr_dict(xlsx_filename, show_warnings)
  
  # Process each flow log csv line:
  with open(flowlog_filename, 'r') as flowlogfile:
    csvreader = csv.DictReader(flowlogfile)
    next(csvreader, None) # Skip header
    total_count = 0
    failed_count = 0
    success_count = 0
    unmatched_count = 0
    for row in csvreader:
      total_count += 1
      src = row['src']
      dst = row['dest']
      numpackets = row['numpackets']

      try:
        src_association = info[src]['associate_with']
        src_name = info[src]['name']
        src_vpc_id = info[src]['vpc_id']
        src_cidr = info[src]['cidr']

        dst_propagations = info[dst]['propagate_to']
        dst_name = info[dst]['name']
        dst_vpc_id = info[dst]['vpc_id']
        dst_cidr = info[dst]['cidr']
      except KeyError as e:
        unmatched_count += 1
        if show_warnings:
          result = f'Warning: no entry for {e}'
          warning_strings.add(result)
          continue

      # Check if src can reach dest:
      if src_association not in dst_propagations:
        failed_count += 1
        result = f'{src_cidr} (src name: {src_name}, src id: {src_vpc_id}) cannot communicate with {dst_cidr} (dst name: {dst_name}, dst id: {dst_vpc_id}) , because the src association {src_association} is not in the dest propagations {dst_propagations}'
        failed_strings.add(result)
      else:
        success_count += 1

    # Finally print the summary:
    for result in failed_strings:
      print(result)

    if show_warnings:
      for result in warning_strings:
        print(result)

    print('')
    print(f'Total processed rows: {total_count}')
    print(f'Unmatched rows: {unmatched_count}')
    print(f'Successful rows: {success_count}')
    print(f'Failed rows: {failed_count}')
    print(f'Deduplicated failed rows: {len(failed_strings)}')

    if unmatched_count > 0 and not show_warnings:
      print('\nRepeat the command with --show-warnings to see unmatched entries')


if __name__ == '__main__':
  main()
